import time
import pandas as pd
import yfinance as yf
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter

# Select the date (adjust the date as needed)
date = '07/05/2024'  # format mm/dd/yyyy
date_format = date.replace('/', '_')

# Set up Chrome options
chrome_options = Options()
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-http2")
chrome_options.add_argument("--disable-web-security")
chrome_options.add_argument("--ignore-certificate-errors")
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--incognito")
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")

# Specify the path to chromedriver
chrome_driver_path = './chromedriver.exe'  # Adjust this path if necessary

# Initialize the Chrome driver
driver = webdriver.Chrome(service=Service(chrome_driver_path), options=chrome_options)
wait = WebDriverWait(driver, 20)

# Function to scroll to an element and ensure it is in the center of the screen
def scroll_to_element(element):
    driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});", element)

def close_popup(element, message):
    close_button = driver.find_element(By.XPATH, element)
    scroll_to_element(close_button)
    close_button.click()
    print(message)

# Function to handle popups
def handle_popups():
    try:
        close_popup('//*[@id="onetrust-accept-btn-handler"]', 'Accepted cookies')
    except Exception as e:
        print("Cookies popup not found or could not click it:", e)
    try:
        close_popup('//*[@id="exp_6f1b787c-5e88-499e-941c-8a18734f8e2d"]/div.exp-ui/div/div[3]/div', 'Closed video popup 1')
    except Exception as e:
        print("Video popup not found or could not click it:", e)
    try:
        close_popup('//*[@id="exp_ba2915bc-100d-4dad-b24f-643fa38b972b"]/div[2]/div/div[3]/div/svg/path', 'Closed video popup 2')
    except Exception as e:
        print("Video popup not found or could not click it:", e)
    try:
        close_popup('//*[@id="exp_ba2915bc-100d-4dad-b24f-643fa38b972b"]/div[2]/div/div[3]/div/svg', 'Closed video popup 3')
    except Exception as e:
        print("Video popup not found or could not click it:", e)

# Retry logic for navigation
retries = 3
for attempt in range(retries):
    try:
        # Navigate to the website
        url = 'https://www.nasdaq.com/market-activity/dividends'
        driver.get(url)
        print("Navigated to the website.")
        time.sleep(3)  # Wait 3 seconds after navigating to the URL
        break
    except Exception as e:
        print(f"Attempt {attempt + 1} failed: {e}")
        if attempt == retries - 1:
            raise
        time.sleep(1.5)  # Wait for 1.5 seconds before retrying

# Wait for the select date button to be clickable and click it
try:
    select_date_button = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/main/div[2]/div[2]/div[2]/div/div[2]/div/div[3]/div[3]/div[2]/button')))
    scroll_to_element(select_date_button)
    select_date_button.click()
    print("Clicked the select date button.")
except Exception as e:
    print("Select date button not clickable or not found:", e)

# Wait for and clear the date input, then type the new date
try:
    date_input = wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[2]/div/main/div[2]/div[2]/div[2]/div/div[2]/div/div[3]/div[3]/div[2]/div/input')))
    date_input.send_keys(Keys.CONTROL + "a")
    date_input.send_keys(Keys.DELETE)
    print("Cleared the date input.")
    time.sleep(0.5)  # Wait to ensure the input is cleared
    date_input.send_keys(date)
    print("Typed the new date.")
except Exception as e:
    print("Date input not found or not interactable:", e)

# Wait for the apply button to be clickable and click it
try:
    apply_button = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/main/div[2]/div[2]/div[2]/div/div[2]/div/div[3]/div[3]/div[2]/div/button[2]')))
    scroll_to_element(apply_button)
    apply_button.click()
    print("Clicked the apply button.")
except Exception as e:
    print("Apply button not clickable or not found:", e)

time.sleep(1.5)  # Wait 1.5 seconds after pressing the apply button
print("Page loaded.")

# Handle popups once at the beginning
handle_popups()
time.sleep(15)

# Extract all columns from the table
columns = ["Symbol", "Name", "Ex-Dividend Date", "Payment Date", "Record Date", "Dividend", "Indicated Annual Dividend", "Announcement"]
additional_columns = ["open", "dayHigh", "dayLow", "volume", "averageVolume", "country", "marketCap", "recommendationKey", "recommendationMean"]
all_columns = columns + additional_columns + ["Investment", "Payout"]

def extract_table_data():
    for attempt in range(4):
        try:
            table_element = driver.find_element(By.XPATH, '/html/body/div[2]/div/main/div[2]/div[2]/div[2]/div/div[2]/div/div[3]/div[5]/div[1]/div/table')
            table_html = table_element.get_attribute('outerHTML')
            soup = BeautifulSoup(table_html, 'lxml')
            table = soup.find('table')
            return pd.read_html(str(table))[0]
        except Exception as e:
            if attempt == 3:
                print(f"Failed to extract table data on page {page_num} after 4 attempts: {e}")
            else:
                print(f"Attempt {attempt + 1} to extract table data failed: {e}")
            time.sleep(1)
    return pd.DataFrame()

# DataFrame to hold all data
all_data = pd.DataFrame(columns=all_columns)

page_num = 1
while True:
    print(f"Extracting data from page {page_num}...")
    data = extract_table_data()
    if data.empty:
        print("No data found on the current page.")
        break

    # Filter out tickers that contain any symbols
    data = data[data['Symbol'].apply(str.isalpha)]
    
    # Process each row to get additional data using yfinance
    processed_data = []
    for _, row_data in data.iterrows():
        symbol = row_data[0]
        row_data = row_data.tolist()
        try:
            stock = yf.Ticker(symbol)
            info = stock.info
            row_data.append(info.get("open", "N/A"))
            row_data.append(info.get("dayHigh", "N/A"))
            row_data.append(info.get("dayLow", "N/A"))
            row_data.append(info.get("volume", "N/A"))
            row_data.append(info.get("averageVolume", "N/A"))
            row_data.append(info.get("country", "N/A"))
            row_data.append(info.get("marketCap", "N/A"))
            row_data.append(info.get("recommendationKey", "N/A"))
            row_data.append(info.get("recommendationMean", "N/A"))
            row_data.append(30000)  # Investment
            row_data.append("N/A")  # Placeholder for Payout calculation
        except Exception as e:
            if '404' in str(e):
                print(f"404 Error for symbol {symbol}: {e}")
            else:
                print(f"Failed to fetch additional data for {symbol}: {e}")
            row_data += ["Not Found"] * (len(additional_columns) + 2)
        processed_data.append(row_data)

    all_data = pd.concat([all_data, pd.DataFrame(processed_data, columns=all_columns)], ignore_index=True)
    print(f"Extracted data from page {page_num}.")

    handle_popups()  # Handle popups after extracting data from each page
    time.sleep(2)  # Wait for 2 seconds to handle popups

    try:
        next_button = driver.find_element(By.XPATH, '/html/body/div[2]/div/main/div[2]/div[2]/div[2]/div/div[2]/div/div[3]/div[6]/button[2]')
        if next_button.is_displayed() and next_button.is_enabled():
            scroll_to_element(next_button)
            next_button.click()
            print("Clicked the next button.")
            page_num += 1
            time.sleep(5)  # Wait for the next page to load
        else:
            print("Next button is not displayed or not enabled.")
            break
    except Exception as e:
        print("No more pages or failed to navigate to next page:", e)
        break

# Perform the Payout calculation
all_data['Payout'] = all_data.apply(lambda row: row['Investment'] / row['open'] * row['Dividend'] if row['open'] != "N/A" and row['Dividend'] != "N/A" else "N/A", axis=1)

# Save the combined data to an Excel file
output_file = f'dividend_data_{date_format}.xlsx'
all_data.to_excel(output_file, index=False)

# Add sort and filter functionality to the Excel file
wb = load_workbook(output_file)
ws = wb.active

# Apply autofilter to all columns
ws.auto_filter.ref = ws.dimensions

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

wb.save(output_file)
wb.close()

print(f"Saved all data to {output_file}.")

driver.quit()
print("Driver quit.")
